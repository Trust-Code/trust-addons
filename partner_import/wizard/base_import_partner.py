# -*- coding: utf-8 -*-
# © 2016 Danimar Ribeiro, Trustcode
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl.html).

import os
import re
import base64
from datetime import datetime
from openpyxl import load_workbook
from openerp import api, fields, models


class baseImportPartner(models.TransientModel):
    _name = 'base.import.partner'
    _description = 'Partner import'

    xls_import = fields.Binary(string="Arquivo a importar", required=True)

    def _format_cpf_cnpj(self, cnpj_cpf):
        val = re.sub('[^0-9]', '', cnpj_cpf)
        if len(val) == 11:
            return "%s.%s.%s-%s" % (val[0:3], val[3:6], val[6:9],
                                    val[9:11])
        elif len(val) == 14:
            return "%s.%s.%s/%s-%s" % (val[0:2], val[2:5], val[5:8],
                                       val[8:12], val[12:14])
        return cnpj_cpf

    def _format_cep(self, cep):
        cep = str(cep)
        if len(cep) == 8:
            return "%s-%s" % (cep[0:5], cep[5:8])
        return cep

    def _search_city(self, state_id, city):
        city = self.env['l10n_br_base.city'].search(
            [('name', '=ilike', city), ('state_id', '=', state_id)], limit=1
        )
        return city.id or None

    def _search_state(self, state):
        state = self.env['res.country.state'].search(
            [('code', '=', state), ('country_id.code', '=', 'BR')], limit=1)
        return state.id or None

    def _next_phone(self, phones, index):
        count = 0
        for item in phones:
            if item:
                if count == index:
                    return item
                count += 1
        return None

    @api.multi
    def action_import_partners(self):
        xlsx = base64.b64decode(self.xls_import)
        tmp = '/tmp/facilitcred/import/planilha.xlsx'
        try:
            os.makedirs(os.path.dirname(tmp))
        except:
            pass
        with open(tmp, 'w') as arquivo:
            arquivo.write(xlsx)

        wb = load_workbook(filename=tmp, read_only=True)
        if len(wb.sheetnames) > 0:
            ws = wb[wb.sheetnames[0]]

        state_env = self.env['res.country.state']
        partner_env = self.env['res.partner']

        indice = 0
        for row in ws.rows:
            valores = []
            for cell in row:
                valores.append(cell.value)
            if indice == 0:
                indice += 1
                continue
            raw_cpf = valores[1]
            temp_cpf = ''
            for i in raw_cpf:
                try:
                    int(i)
                    temp_cpf += str(i)
                except:
                    pass
            true_cpf = int(temp_cpf)
            cpf = self._format_cpf_cnpj('%0.11d' % (true_cpf or 0))

            partner = self.env['res.partner'].search([('cnpj_cpf', '=', cpf)])
            telefones = [valores[10], valores[12], valores[17], valores[18],
                         valores[19], valores[20]]
            state = self._search_state(valores[8])
            birth_date = valores[11]
            if isinstance(birth_date, basestring):
                birth_date = datetime.strptime(birth_date, '%d/%m/%Y')
            vals = {
                'name': valores[0],
                'cnpj_cpf': cpf,
                'street': valores[2],
                'street2': valores[3],
                'number': valores[4],
                'zip': self._format_cep(valores[5]),
                'district': valores[6],
                'l10n_br_city_id': self._search_city(state, valores[7]),
                'state_id': state,
                'phone': self._next_phone(telefones, 0),
                'birthdate_date': birth_date,
                'mobile': self._next_phone(telefones, 1),
                'x_num_beneficio': valores[13],
                'x_phone1': self._next_phone(telefones, 2),
                'fax': self._next_phone(telefones, 3),
                'x_tipo': valores[21],
            }
            vals = {k: v for k, v in vals.items() if v}
            if partner:
                if partner.x_num_beneficio != \
                   vals.get('x_num_beneficio', False) and \
                   not partner.x_num_beneficio:
                    partner.x_num_beneficio = vals['x_num_beneficio']
                    vals['x_num_beneficio'] = False
                elif partner.x_num_beneficio_2 != \
                    vals.get('x_num_beneficio', False) and \
                        not partner.x_num_beneficio_2:
                    partner.x_num_beneficio_2 = vals['x_num_beneficio']
                    vals['x_num_beneficio'] = False
                elif partner.x_num_beneficio_3 != \
                    vals.get('x_num_beneficio', False) and \
                        not partner.x_num_beneficio_3:
                    partner.x_num_beneficio_3 = vals['x_num_beneficio']
                    vals['x_num_beneficio'] = False
                vals = {k: v for k, v in vals.items() if v}
                partner.write(vals)
            else:
                self.env['res.partner'].create(vals)
            if not partner.l10n_br_city_id and partner.zip and \
               len(partner.zip) == 9:
                try:
                    partner.zip_search()
                except:
                    pass

            indice += 1
